import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

KLANTEN_HEADERS = [
    "klant_id", "voornaam", "achternaam", "adres", "telefoon",
    "rooster", "porties", "vast_toetje", "bezorgwijze",
    "vaste_bezorgtijd", "bezorg_opmerkingen", "dieetwensen", "actief",
]

KLANTEN = [
    [160, "Gerrit", "Bazuin", "", "", "vr", 1, "nee", "ter_plaatse", "17:30", "", "GEEN chili con carne; geen prei", "ja"],
    [182, "Marrie", "Scholten", "Neptunus Plein 81", "", "", 1, "nee", "bezorgen", "17:00", "Rooster nog vaststellen", "GEEN PAPRIKA", "ja"],
    [216, "Greet", "Westerdijk-Vos", "Nijlandstraat 102", "0592-482214", "ma,di,wo,do,vr,za", 1, "ja", "bezorgen", "12:00", "", "weinig aardappels; veel groente", "ja"],
    [218, "Ferdinand", "Bloemen", "Citadel 21", "", "vr", 1, "nee", "ter_plaatse", "17:30", "Bellen voor bezorging", "", "ja"],
    [224, "Grietje", "Vrieling", "Poststraat 20", "", "", 1, "nee", "bezorgen", "17:00", "Rooster nog vaststellen", "", "ja"],
    [230, "Kitty", "Boogaard", "Dotterbloemstraat 127", "06-58999597", "do,vr", 1, "nee", "bezorgen", "17:00", "", "Gaar eten; geen velletjes; geen salade", "ja"],
    [234, "Marga", "Radix", "Smetanalaan 264", "", "ma,wo,vr", 1, "nee", "bezorgen", "17:00", "", "", "ja"],
    [242, "Geertje", "Speelman", "Jan Fabriciusstraat 113", "+31681479920", "", 1, "nee", "bezorgen", "12:00", "Rooster nog vaststellen", "", "ja"],
    [246, "Hadassa", "Loopuit", "Zaagmolen 2", "", "ma,di,wo,do,vr,za", 1, "nee", "bezorgen", "17:15", "", "Vegetarisch; wel zalm en kabeljauw; GEEN garnalen; GEEN PESTO", "ja"],
    [264, "Anneke", "van der Pal", "", "", "vr", 1, "nee", "ter_plaatse", "17:30", "", "", "ja"],
    [283, "Peter", "Westerveld", "Van Goghstraat 34, 9403 CK", "", "", 1, "nee", "bezorgen", "17:00", "Gestopt", "Geen salade; GEEN cappucijners; geen bruine bonen; geen doperwten; geen rookworst", "nee"],
    [291, "Fam.", "De Boer", "Zwartewaterseweg 64", "0592-341904", "ma,wo,vr", 2, "nee", "bezorgen", "12:00", "", "GEEN olijven; Mr. WEL vis; Mevr. geen vis", "ja"],
    [293, "Eppie", "Bartels", "", "", "vr", 1, "nee", "ter_plaatse", "17:30", "", "", "ja"],
    [320, "Jorrit en Rachelle", "", "", "", "", 1, "nee", "ophalen", "18:00", "", "", "nee"],
    [325, "Greet", "Havinga", "Citadel 70", "06-10724308", "di", 1, "nee", "bezorgen", "18:00", "Alleen in even weken — oneven weken via Afwijkingen annuleren", "Niet te vet", "ja"],
    [9001, "Ed", "van Dijk", "Groningerdwartstraat 26", "06-10173742", "", 1, "nee", "bezorgen", "12:00", "Gestopt", "Geen snert; geen bonensoep; geen boerenkool", "nee"],
    [9002, "Roelie", "Ma Koops", "Vaart Noordzijde 74 E", "", "ma,wo,vr", 1, "nee", "bezorgen", "17:00", "", "Liever geen vis", "ja"],
    [9003, "Kor", "Bruggink", "Troelstralaan 66", "", "ma,di,wo,do,vr", 1, "nee", "bezorgen", "17:30", "", "", "ja"],
    [9004, "Johanna", "IJszenga", "Mondriaanstraat 7, 9403 BN", "", "ma,di,wo,do,vr,za,zo", 1, "nee", "bezorgen", "17:10", "", "Geen spek; geen kool; VEZELRIJK", "ja"],
    [9005, "Rosanne", "Van der Zwet", "Stelling", "", "ma,di,wo,do,vr", 1, "nee", "ophalen", "17:00", "", "", "ja"],
    [9006, "Dolf", "Zuidhof", "Smetanalaan 43a", "", "", 1, "nee", "bezorgen", "18:30", "", "", "nee"],
]

AFWIJKINGEN_HEADERS = [
    "datum", "weekdag", "klant_id", "type",
    "porties", "toetje", "tijd", "bezorgwijze", "bezorger", "notitie",
]

AFWIJKINGEN = [
    ["", "wo", 234, "wijziging", "", "", "18:00", "", "", "Vaste woensdag-tijd"],
    ["", "do", 9005, "wijziging", "", "", "18:30", "", "", "Donderdag pas na 18:30"],
    ["", "zo", 9004, "wijziging", "", "", "", "", "Iris", "Vaste zondagsbezorger"],
]

wb = Workbook()

ws1 = wb.active
ws1.title = "Klanten"
ws1.append(KLANTEN_HEADERS)
for row in KLANTEN:
    ws1.append(row)

header_font = Font(bold=True)
header_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
for cell in ws1[1]:
    cell.font = header_font
    cell.fill = header_fill
ws1.freeze_panes = "A2"
widths = {"A": 9, "B": 18, "C": 16, "D": 28, "E": 14, "F": 20, "G": 8, "H": 11, "I": 13, "J": 11, "K": 32, "L": 38, "M": 7}
for col, w in widths.items():
    ws1.column_dimensions[col].width = w

ws2 = wb.create_sheet("Afwijkingen")
ws2.append(AFWIJKINGEN_HEADERS)
for row in AFWIJKINGEN:
    ws2.append(row)
for cell in ws2[1]:
    cell.font = header_font
    cell.fill = header_fill
ws2.freeze_panes = "A2"
widths2 = {"A": 12, "B": 9, "C": 9, "D": 12, "E": 8, "F": 7, "G": 7, "H": 13, "I": 12, "J": 40}
for col, w in widths2.items():
    ws2.column_dimensions[col].width = w

out_path = "/tmp/bezorgservice.xlsx"
wb.save(out_path)

with open(out_path, "rb") as f:
    b64 = base64.b64encode(f.read()).decode()

with open("/tmp/bezorgservice.b64", "w") as f:
    f.write(b64)

print(f"base64 chars: {len(b64)}")
