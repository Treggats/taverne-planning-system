import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

MEDEWERKERS_HEADERS = ["naam", "type", "actief"]

# Geseed uit de werkrooster-tabs van de weekplanningen. Namen/types kan Antje
# bijwerken via de medewerker-actie of direct in de sheet.
MEDEWERKERS = [
    ["Antje", "vast", "ja"],
    ["Sus", "vast", "ja"],
    ["Tonko", "vast", "ja"],
    ["Iris", "vast", "ja"],
    ["Jeffrey", "vast", "ja"],
    ["Renate", "vast", "ja"],
    ["Arjan", "vast", "ja"],
    ["Elizabeth", "vast", "ja"],
    ["Sieme", "vast", "ja"],
    ["Tessa", "vast", "ja"],
    ["Cas", "vast", "ja"],
    ["Roelof", "vast", "ja"],
    ["Noreen", "vast", "ja"],
    ["Sara", "vast", "ja"],
    ["Hanah", "vast", "ja"],
    ["Laura", "stagiair", "ja"],
    ["Faye", "stagiair", "ja"],
    ["Tijn", "stagiair", "ja"],
    ["Amar", "stagiair", "ja"],
    ["Frans Jozef", "vrijwilliger", "ja"],
    ["Peter", "vrijwilliger", "ja"],
    ["Marye", "vrijwilliger", "ja"],
    ["Rosanne", "vrijwilliger", "ja"],
    ["Heidi", "vrijwilliger", "ja"],
    ["Damaris", "vrijwilliger", "nee"],
]

# Snapshot per week: één rij per persoon per werkdag. Start leeg.
DIENSTEN_HEADERS = ["week", "weekdag", "naam", "begin", "eind", "notitie"]

wb = Workbook()

ws1 = wb.active
ws1.title = "Medewerkers"
ws1.append(MEDEWERKERS_HEADERS)
for row in MEDEWERKERS:
    ws1.append(row)

header_font = Font(bold=True)
header_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
for cell in ws1[1]:
    cell.font = header_font
    cell.fill = header_fill
ws1.freeze_panes = "A2"
for col, w in {"A": 18, "B": 14, "C": 8}.items():
    ws1.column_dimensions[col].width = w

ws2 = wb.create_sheet("Diensten")
ws2.append(DIENSTEN_HEADERS)
for cell in ws2[1]:
    cell.font = header_font
    cell.fill = header_fill
ws2.freeze_panes = "A2"
for col, w in {"A": 10, "B": 9, "C": 18, "D": 8, "E": 8, "F": 30}.items():
    ws2.column_dimensions[col].width = w

out_path = "/tmp/werkrooster.xlsx"
wb.save(out_path)

with open(out_path, "rb") as f:
    b64 = base64.b64encode(f.read()).decode()

with open("/tmp/werkrooster.b64", "w") as f:
    f.write(b64)

print(f"base64 chars: {len(b64)}")
